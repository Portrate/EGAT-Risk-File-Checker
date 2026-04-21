import io
import json
import os
import sys
import urllib.parse
from datetime import datetime

import httpx
import openpyxl
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import ValidationError

from checker.models import ChecklistSection
from checker.orchestrator import MODEL, analyze_with_llm
from checker.pdf_extractor import extract_text_from_bytes

# When frozen by PyInstaller, files are extracted to sys._MEIPASS
BASE_DIR = sys._MEIPASS if getattr(sys, "frozen", False) else os.path.dirname(os.path.abspath(__file__))

app = FastAPI()

app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))


@app.get("/")
def index(request: Request):
    return templates.TemplateResponse(request, "index.html")


@app.post("/analyze")
async def analyze(
    file: UploadFile = File(...),
    checklist: str = Form(...),
):
    pdf_bytes = await file.read()
    document_text = extract_text_from_bytes(pdf_bytes)

    try:
        raw = json.loads(checklist)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=422, detail=f"checklist JSON ไม่ถูกต้อง: {e}")

    try:
        sections = [ChecklistSection.model_validate(s) for s in raw]
    except ValidationError as e:
        raise HTTPException(status_code=422, detail=e.errors())

    checklist_data = [{"title": s.title, "items": [{"name": i.name, "score": i.score} for i in s.items]} for s in sections]

    try:
        llm_result = await analyze_with_llm(document_text, checklist_data)
    except httpx.ConnectError:
        raise HTTPException(
            status_code=502,
            detail="ไม่สามารถเชื่อมต่อ Ollama ได้ — โปรดตรวจสอบว่า Ollama กำลังรันอยู่ที่ localhost:11434",
        )
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=502, detail=f"Ollama ตอบกลับด้วย error: {e.response.status_code}")
    except (json.JSONDecodeError, KeyError, ValueError) as e:
        raise HTTPException(status_code=500, detail=f"แปลงผลลัพธ์จาก LLM ไม่ได้: {e}")

    results = llm_result.get("results", [])

    total = 0
    passed = 0
    total_score = 0.0
    passed_score = 0.0
    for section in results:
        for item in section.get("items", []):
            total += 1
            item_score = float(item.get("score", 0.0))
            total_score += item_score
            if item.get("status") == "pass":
                passed += 1
                passed_score += item_score

    score = round(passed / total * 100) if total > 0 else 0

    return {
        "summary": {
            "total": total,
            "passed": passed,
            "failed": total - passed,
            "total_score": total_score,
            "passed_score": passed_score
        },
        "similarity_score": score,
        "results": results,
    }


@app.post("/export/excel")
async def export_excel(payload: dict):
    filename = payload.get("filename", "document")
    results_data = payload.get("results", [])
    summary = payload.get("summary", {})
    score = payload.get("similarity_score", 0)

    wb = openpyxl.Workbook()

    # ---- Sheet 1: Summary ----
    ws1 = wb.active
    ws1.title = "สรุปผล"

    header_fill = PatternFill("solid", fgColor="1A73E8")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 30

    summary_rows = [
        ("ชื่อไฟล์", filename),
        ("วันที่ตรวจสอบ", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("จำนวนรายการทั้งหมด", summary.get("total", 0)),
        ("ผ่าน (FOUND)", summary.get("passed", 0)),
        ("ไม่ผ่าน (MISSING)", summary.get("failed", 0)),
        ("คะแนนที่ได้", f"{summary.get('passed_score', 0)} / {summary.get('total_score', 0)}"),
        ("คะแนน (%)", f"{score}%"),
    ]

    for label, value in summary_rows:
        row = ws1.append([label, value])

    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row):
        row[0].font = Font(bold=True)
        row[0].alignment = left
        row[1].alignment = left

    # ---- Sheet 2: Detailed results ----
    ws2 = wb.create_sheet("ผลการตรวจสอบ")

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 36
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 40
    ws2.column_dimensions["G"].width = 40

    headers = ["หัวข้อหลัก", "รายการย่อย", "คะแนนเต็ม", "คะแนนที่ได้", "ผล", "เหตุผล", "หลักฐาน"]
    ws2.append(headers)
    header_row = ws2[1]
    for cell in header_row:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center

    ws2.row_dimensions[1].height = 28

    pass_fill = PatternFill("solid", fgColor="E6F4EA")
    fail_fill = PatternFill("solid", fgColor="FCE8E6")
    pass_font = Font(color="137333", bold=True)
    fail_font = Font(color="C5221F", bold=True)
    score_pass_font = Font(color="137333", bold=True)
    score_fail_font = Font(color="C5221F", bold=True)

    for section in results_data:
        section_title = section.get("section", "")
        for item in section.get("items", []):
            status = item.get("status", "")
            is_pass = status == "pass"
            max_score = item.get("score", 0.0)
            earned_score = max_score if is_pass else 0.0
            ws2.append([
                section_title,
                item.get("requirement", ""),
                max_score,
                earned_score,
                "FOUND" if is_pass else "MISSING",
                item.get("reasoning", ""),
                item.get("evidence") or "",
            ])
            data_row = ws2[ws2.max_row]
            row_fill = pass_fill if is_pass else fail_fill
            for cell in data_row:
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            data_row[2].alignment = center
            data_row[3].font = score_pass_font if is_pass else score_fail_font
            data_row[3].alignment = center
            data_row[4].font = pass_font if is_pass else fail_font
            data_row[4].alignment = center
            ws2.row_dimensions[ws2.max_row].height = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_base = filename.removesuffix(".pdf").replace(" ", "_")
    out_name = f"ผลตรวจสอบ_{safe_base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_name = urllib.parse.quote(out_name)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"},
    )


@app.get("/health")
async def health():
    """Check FastAPI + Ollama availability."""
    ollama_ok = False
    try:
        async with httpx.AsyncClient(timeout=5) as client:
            r = await client.get("http://localhost:11434/api/tags")
            ollama_ok = r.status_code == 200
    except Exception:
        pass

    return {
        "status": "ok",
        "ollama": "reachable" if ollama_ok else "unreachable",
        "model": MODEL,
    }
