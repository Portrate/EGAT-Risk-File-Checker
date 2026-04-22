import io
import urllib.parse
from datetime import datetime

import openpyxl
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill


def build_excel_response(filename: str, results_data: list, summary: dict, score: int) -> StreamingResponse:
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Summary ----
    ws1 = wb.active
    ws1.title = "สรุปผล"

    header_fill = PatternFill("solid", fgColor="1A73E8")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 30

    summary_rows = [
        ("ชื่อไฟล์", filename),
        ("วันที่ตรวจสอบ", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("จำนวนรายการทั้งหมด", summary.get("total", 0)),
        ("ผ่าน (FOUND)", summary.get("passed", 0)),
        ("ไม่ผ่าน (MISSING)", summary.get("failed", 0)),
        ("คะแนนที่ได้", f"{summary.get('passed_score', 0)} / {summary.get('total_score', 0)}"),
        ("เปอร์เซ็นต์ความครบถ้วน", f"{score}%"),
    ]

    for label, value in summary_rows:
        ws1.append([label, value])

    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row):
        row[0].font = Font(bold=True)
        row[0].alignment = left
        row[1].alignment = left

    # ---- Sheet 2: Detailed results ----
    ws2 = wb.create_sheet("ผลการตรวจสอบ")

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 36
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 40
    ws2.column_dimensions["G"].width = 40

    headers = ["หัวข้อหลัก", "รายการย่อย", "คะแนนเต็ม", "คะแนนที่ได้", "ผล", "เหตุผล", "หลักฐาน"]
    ws2.append(headers)
    header_row = ws2[1]
    for cell in header_row:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center

    ws2.row_dimensions[1].height = 28

    pass_fill = PatternFill("solid", fgColor="E6F4EA")
    fail_fill = PatternFill("solid", fgColor="FCE8E6")
    pass_font = Font(color="137333", bold=True)
    fail_font = Font(color="C5221F", bold=True)
    score_pass_font = Font(color="137333", bold=True)
    score_fail_font = Font(color="C5221F", bold=True)

    for section in results_data:
        section_title = section.get("section", "")
        for item in section.get("items", []):
            status = item.get("status", "")
            is_pass = status == "pass"
            max_score = item.get("score", 0.0)
            earned_score = max_score if is_pass else 0.0
            ws2.append([
                section_title,
                item.get("requirement", ""),
                max_score,
                earned_score,
                "FOUND" if is_pass else "MISSING",
                item.get("reasoning", ""),
                item.get("evidence") or "",
            ])
            data_row = ws2[ws2.max_row]
            row_fill = pass_fill if is_pass else fail_fill
            for cell in data_row:
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            data_row[2].alignment = center
            data_row[3].font = score_pass_font if is_pass else score_fail_font
            data_row[3].alignment = center
            data_row[4].font = pass_font if is_pass else fail_font
            data_row[4].alignment = center
            ws2.row_dimensions[ws2.max_row].height = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_base = filename.removesuffix(".pdf").replace(" ", "_")
    out_name = f"ผลตรวจสอบ_{safe_base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_name = urllib.parse.quote(out_name)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"},
    )
